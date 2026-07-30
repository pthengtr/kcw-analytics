"""Monthly multi-account bank statement Excel report.

Reads Supabase bank.statement_lines (including live match_status / match_reason),
enriches with raw_json channel/detail, and writes one workbook per month
(one sheet per account) under Drive 04_outputs/04_Bank_Statement_Report.
"""

from __future__ import annotations

import json
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from src.kcw import paths
from src.kcw.tar import supabase_db_url

BANGKOK = ZoneInfo("Asia/Bangkok")

COMPANY_NAME = "บริษัท เกียรติชัยอะไหล่ยนต์ 2007 จำกัด (สำนักงานใหญ่)"
COMPANY_ADDRESS = "305 หมู่ 1 ต.ชุมแสง อ.วังจันทร์ จ.ระยอง 21210"
TAX_ID = "0215560000262"

COLUMN_ORDER = [
    "ลำดับ",
    "วันที่",
    "วันที่มีผล",
    "รายการ",
    "ช่องทาง",
    "รายละเอียด",
    "อ้างอิง",
    "ถอนเงิน",
    "ฝากเงิน",
    "ยอดคงเหลือ",
    "แหล่งไฟล์",
    "เหตุผลการจับคู่",
    "สถานะจับคู่",
]

MONEY_COLS = {"ถอนเงิน", "ฝากเงิน", "ยอดคงเหลือ"}
DATE_COLS = {"วันที่", "วันที่มีผล"}
# Soft amber for rows that are not done yet.
DONE_MATCH_STATUSES = frozenset({"matched", "manual", "resolved"})
WARNING_FILL = PatternFill(start_color="FFF2CC", fill_type="solid")

CHANNEL_KEYS = (
    "ช่องทาง",
    "channel",
    "Channel",
    "CHANNEL",
    "Teller Id",
    "Teller ID",
    "Init Br.",
)
# Exact keys only — do NOT substring-match "รายการ" (hits "เวลา/วันที่ ทำรายการ").
TXN_TYPE_KEYS = ("รายการ", "Transaction Code", "TRANSACTION CODE")
DETAIL_KEYS = (
    "รายละเอียด",
    "Description",
    "DESCRIPTION",
    "Particular",
    "PARTICULAR",
)
TIME_KEYS = (
    "เวลา/วันที่ ทำรายการ",
    "เวลา/ วันที่มีผล",
    "เวลา",
    "TIME",
    "Time",
)


def reporting_year_month(as_of: date | None = None) -> tuple[int, int]:
    """VAT-style reporting month: Asia/Bangkok today − 10 days."""
    if as_of is None:
        as_of = datetime.now(BANGKOK).date()
    target = as_of - timedelta(days=10)
    return target.year, target.month


def output_dir() -> Path:
    return paths.ensure_dir(paths.outputs_dir() / "04_Bank_Statement_Report")


def output_path(year: int, month: int) -> Path:
    return output_dir() / f"bank_statement_report_{year}_{month:02d}.xlsx"


def _parse_raw_json(value: Any) -> dict:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return {}
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return {}
        try:
            parsed = json.loads(text)
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            return {}
    return {}


def _pick_raw_value(raw: dict, keys: tuple[str, ...], *, exact_only: bool = False) -> str:
    if not raw:
        return ""
    # Exact key first
    for key in keys:
        if key in raw and raw[key] is not None:
            text = str(raw[key]).strip()
            if text and text.lower() not in {"nan", "none", "<na>"}:
                return text
    # Case-insensitive exact key
    lowered = {str(k).strip().lower(): v for k, v in raw.items()}
    for key in keys:
        lk = key.lower()
        if lk in lowered and lowered[lk] is not None:
            text = str(lowered[lk]).strip()
            if text and text.lower() not in {"nan", "none", "<na>"}:
                return text
    if exact_only:
        return ""
    # Substring fallback (avoid matching time cols that contain "รายการ")
    for rk, rv in raw.items():
        rk_l = str(rk).strip().lower()
        if "เวลา" in rk_l or "time" in rk_l:
            continue
        for key in keys:
            if key.lower() in rk_l and rv is not None:
                text = str(rv).strip()
                if text and text.lower() not in {"nan", "none", "<na>"}:
                    return text
    return ""


def extract_raw_fields(raw_json: Any) -> tuple[str, str, str, str]:
    """Return (txn_type, channel, detail, time_str) from statement raw_json."""
    raw = _parse_raw_json(raw_json)
    txn_type = _pick_raw_value(raw, TXN_TYPE_KEYS, exact_only=True)
    channel = _pick_raw_value(raw, CHANNEL_KEYS, exact_only=True)
    detail = _pick_raw_value(raw, DETAIL_KEYS, exact_only=True)
    time_str = _pick_raw_value(raw, TIME_KEYS, exact_only=True)
    return txn_type, channel, detail, time_str


def _looks_like_time(value: Any) -> bool:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return False
    text = str(value).strip()
    return bool(re.fullmatch(r"\d{1,2}:\d{2}(:\d{2})?", text))


def load_statement_lines(
    conn,
    year: int,
    month: int,
) -> pd.DataFrame:
    """Load bank.statement_lines for calendar month, joined to import filename."""
    sql = """
    SELECT
      l.account_no,
      l.bank_name,
      l.txn_date,
      l.value_date,
      l.description,
      l.bank_reference,
      l.amount,
      l.direction,
      l.debit,
      l.credit,
      l.balance_after,
      l.raw_json,
      l.source_row_number,
      l.source_file_id,
      l.match_status,
      l.match_reason,
      l.match_notes,
      l.matched_ref_type,
      l.matched_ref_id,
      l.match_confidence,
      f.original_filename
    FROM bank.statement_lines l
    LEFT JOIN bank.statement_import_files f ON f.id = l.source_file_id
    WHERE l.txn_date >= %s::date
      AND l.txn_date < %s::date
    ORDER BY l.account_no, l.txn_date, l.value_date NULLS LAST, l.source_row_number NULLS LAST
    """
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1)
    else:
        end = date(year, month + 1, 1)

    with conn.cursor() as cur:
        cur.execute(sql, (start.isoformat(), end.isoformat()))
        cols = [d[0] for d in cur.description]
        rows = cur.fetchall()
    return pd.DataFrame(rows, columns=cols)


def format_match_reason(row: pd.Series) -> str:
    """Build Excel reason text from live Supabase match columns."""
    parts: list[str] = []
    reason = row.get("match_reason")
    notes = row.get("match_notes")
    ref_type = row.get("matched_ref_type")
    ref_id = row.get("matched_ref_id")
    conf = row.get("match_confidence")

    if pd.notna(reason) and str(reason).strip():
        parts.append(str(reason).strip())
    ref_bits = []
    if pd.notna(ref_type) and str(ref_type).strip():
        ref_bits.append(str(ref_type).strip())
    if pd.notna(ref_id) and str(ref_id).strip():
        ref_bits.append(str(ref_id).strip())
    if ref_bits:
        parts.append(" / ".join(ref_bits))
    if pd.notna(notes) and str(notes).strip():
        parts.append(str(notes).strip())
    if pd.notna(conf) and str(conf).strip() != "":
        try:
            parts.append(f"confidence={float(conf):.2f}")
        except (TypeError, ValueError):
            parts.append(f"confidence={conf}")
    return "\n".join(parts) if parts else ""


def enrich_statement_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Add channel/detail from raw_json; use live DB match_status/reason."""
    if df.empty:
        return pd.DataFrame(columns=COLUMN_ORDER + ["account_no", "bank_name"])

    out = df.copy()
    txn_types: list[str] = []
    channels: list[str] = []
    details: list[str] = []
    times: list[str] = []
    for raw in out.get("raw_json", pd.Series([None] * len(out))):
        txn, ch, det, tm = extract_raw_fields(raw)
        txn_types.append(txn)
        channels.append(ch)
        details.append(det)
        times.append(tm)
    out["ช่องทาง"] = channels
    out["รายละเอียด"] = details

    # Prefer raw_json รายการ; imported description was often the time column
    # (import bug: pattern "รายการ" matched "เวลา/วันที่ ทำรายการ").
    descriptions: list[str] = []
    for txn, desc in zip(txn_types, out.get("description", pd.Series([None] * len(out)))):
        if txn:
            descriptions.append(txn)
        elif _looks_like_time(desc):
            descriptions.append("")
        else:
            descriptions.append(
                ""
                if desc is None or (isinstance(desc, float) and pd.isna(desc))
                else str(desc)
            )
    out["รายการ"] = descriptions

    value_dates: list[Any] = []
    for vd, tm in zip(out.get("value_date", pd.Series([None] * len(out))), times):
        if pd.notna(vd) and not _looks_like_time(vd):
            value_dates.append(vd)
        elif tm:
            value_dates.append(tm)
        else:
            value_dates.append(vd)
    out["_value_display"] = value_dates

    out["ถอนเงิน"] = pd.to_numeric(out.get("debit"), errors="coerce")
    out["ฝากเงิน"] = pd.to_numeric(out.get("credit"), errors="coerce")
    out["ยอดคงเหลือ"] = pd.to_numeric(out.get("balance_after"), errors="coerce")

    # Live match fields from Supabase (not heuristic amount matching)
    if "match_status" in out.columns:
        status = out["match_status"].fillna("").astype(str).str.strip()
        status = status.replace({"": "pending", "None": "pending", "nan": "pending"})
        out["สถานะจับคู่"] = status
    else:
        out["สถานะจับคู่"] = "pending"
    out["เหตุผลการจับคู่"] = out.apply(format_match_reason, axis=1)

    out["อ้างอิง"] = out.get("bank_reference", pd.Series([None] * len(out))).fillna("").astype(str)
    out["แหล่งไฟล์"] = out.get("original_filename", pd.Series([None] * len(out))).fillna("").astype(str)
    out["วันที่"] = pd.to_datetime(out.get("txn_date"), errors="coerce")
    out["วันที่มีผล"] = out["_value_display"]
    out = out.drop(columns=["_value_display"], errors="ignore")

    return out


def sheet_name_for_account(bank_name: Any, account_no: Any) -> str:
    bank = str(bank_name or "").strip() or "BANK"
    acct = str(account_no or "").strip() or "UNKNOWN"
    name = f"{bank}_{acct}"
    # Excel sheet name limits
    name = re.sub(r'[:\\/?*\[\]]', "_", name)
    return name[:31]


def build_account_sheets(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Group enriched rows into per-account report DataFrames."""
    sheets: dict[str, pd.DataFrame] = {}
    if df.empty:
        return sheets

    group_cols = ["bank_name", "account_no"]
    for (bank, acct), group in df.groupby(group_cols, dropna=False, sort=True):
        g = group.copy()
        g = g.sort_values(
            ["วันที่", "วันที่มีผล", "source_row_number"],
            kind="stable",
            na_position="last",
        ).reset_index(drop=True)
        g["ลำดับ"] = range(1, len(g) + 1)

        export_df = pd.DataFrame({c: g[c] if c in g.columns else None for c in COLUMN_ORDER})
        name = sheet_name_for_account(bank, acct)
        # Disambiguate collisions after truncation
        base = name
        i = 2
        while name in sheets:
            suffix = f"_{i}"
            name = (base[: 31 - len(suffix)] + suffix)
            i += 1
        sheets[name] = export_df
    return sheets


def _add_total_row(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    total = {c: "" for c in df.columns}
    total["รายการ"] = "รวม"
    for col in ("ถอนเงิน", "ฝากเงิน"):
        if col in df.columns:
            total[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).sum()
    if "ลำดับ" in total:
        total["ลำดับ"] = ""
    return pd.concat([df, pd.DataFrame([total])], ignore_index=True)


def export_workbook(
    sheets: dict[str, pd.DataFrame],
    out_path: Path,
    *,
    year: int,
    month: int,
    company_name: str = COMPANY_NAME,
    company_address: str = COMPANY_ADDRESS,
    tax_id: str = TAX_ID,
) -> Path:
    """Write multi-sheet workbook with VAT-style openpyxl formatting."""
    wb = Workbook()
    if wb.sheetnames == ["Sheet"]:
        wb.remove(wb["Sheet"])

    title_name = f"รายงานเดินบัญชีธนาคาร ประจำเดือน {month:02d}/{year}"
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="F3DFD2", fill_type="solid")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if not sheets:
        # Still create a placeholder sheet so the file exists
        sheets = {"NO_DATA": pd.DataFrame(columns=COLUMN_ORDER)}

    for sheet_name, df in sheets.items():
        export_df = _add_total_row(df.copy())
        # Replace NA for Excel
        export_df = export_df.where(pd.notna(export_df), None)

        ws = wb.create_sheet(sheet_name)
        last_col = max(len(export_df.columns), 8)
        end_letter = get_column_letter(last_col)

        ws.merge_cells(f"A1:{end_letter}1")
        ws["A1"] = f"{title_name} — {sheet_name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = center

        ws.merge_cells("A3:C3")
        ws["A3"] = "ชื่อสถานประกอบกิจการ"
        ws.merge_cells("D3:F3")
        ws["D3"] = company_name
        ws.merge_cells("A4:C4")
        ws["A4"] = "ที่อยู่สถานประกอบกิจการ"
        ws.merge_cells("D4:F4")
        ws["D4"] = company_address
        ws.merge_cells("A5:C5")
        ws["A5"] = "เลขประจำตัวผู้เสียภาษี"
        ws.merge_cells("D5:F5")
        ws["D5"] = tax_id

        for r in (3, 4, 5):
            ws[f"A{r}"].font = bold
            ws[f"A{r}"].alignment = left
            ws[f"D{r}"].alignment = left

        start_row = 7
        status_col_idx = None
        if "สถานะจับคู่" in export_df.columns:
            status_col_idx = list(export_df.columns).index("สถานะจับคู่") + 1

        for r_idx, row in enumerate(
            dataframe_to_rows(export_df, index=False, header=True), start_row
        ):
            is_header = r_idx == start_row
            status_val = ""
            if not is_header and status_col_idx is not None:
                # row is a sequence matching columns; status at status_col_idx-1
                status_val = str(row[status_col_idx - 1] or "").strip().lower()
            warn_row = (
                (not is_header)
                and bool(status_val)
                and status_val not in DONE_MATCH_STATUSES
            )

            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                if is_header:
                    cell.font = bold
                    cell.fill = header_fill
                    cell.alignment = center
                else:
                    cell.alignment = left
                    if warn_row:
                        cell.fill = WARNING_FILL

        ws.freeze_panes = f"A{start_row + 1}"

        for c_idx, col_name in enumerate(export_df.columns, 1):
            series = export_df[col_name].astype(str).fillna("")
            max_len = max(int(series.map(len).max() if len(series) else 0), len(str(col_name)))
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 2, 60)

        data_end = start_row + len(export_df)
        for c_idx, col_name in enumerate(export_df.columns, 1):
            name = str(col_name).strip()
            if name in MONEY_COLS:
                for r in range(start_row + 1, data_end + 1):
                    ws.cell(row=r, column=c_idx).number_format = "#,##0.00"
            if name in DATE_COLS:
                for r in range(start_row + 1, data_end + 1):
                    ws.cell(row=r, column=c_idx).number_format = "dd/mm/yyyy"

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def write_fixture_sample(out_path: Path | None = None) -> Path:
    """Synthetic multi-account workbook using the same exporter (no Drive/DB)."""
    if out_path is None:
        out_path = paths.log_dir() / "bank_statement_report_fixture_sample.xlsx"
    paths.ensure_dir(out_path.parent)

    year, month = 2026, 6
    rows = pd.DataFrame(
        [
            {
                "account_no": "0393",
                "bank_name": "KBANK",
                "txn_date": date(2026, 6, 2),
                "value_date": date(2026, 6, 2),
                "description": "โอนเงิน",
                "bank_reference": "REF001",
                "debit": 1500.00,
                "credit": None,
                "balance_after": 98500.00,
                "raw_json": {
                    "รายการ": "โอนเงิน",
                    "ช่องทาง": "K BIZ",
                    "รายละเอียด": "โอนไป BBL X2440",
                },
                "original_filename": "KBANK0393_sample.xlsx",
                "source_row_number": 1,
                "match_status": "matched",
                "match_reason": "ใบสำคัญจ่าย",
                "match_notes": "จับคู่กับ PV2606001",
                "matched_ref_type": "voucher",
                "matched_ref_id": "PV2606001",
                "match_confidence": 0.95,
            },
            {
                "account_no": "0393",
                "bank_name": "KBANK",
                "txn_date": date(2026, 6, 3),
                "value_date": date(2026, 6, 3),
                "description": "รับโอน",
                "bank_reference": "REF002",
                "debit": None,
                "credit": 25000.00,
                "balance_after": 123500.00,
                "raw_json": {
                    "รายการ": "รับโอนเงิน",
                    "ช่องทาง": "K BIZ",
                    "รายละเอียด": "รับจากลูกค้า",
                },
                "original_filename": "KBANK0393_sample.xlsx",
                "source_row_number": 2,
                "match_status": "matched",
                "match_reason": "ยอดขายสุทธิ 3TAR",
                "match_notes": "ยอดขายสุทธิรายวันเข้าบัญชี",
                "matched_ref_type": "tar_cntar_net",
                "matched_ref_id": "2026-06-02",
                "match_confidence": 0.95,
            },
            {
                "account_no": "6184",
                "bank_name": "KTB",
                "txn_date": date(2026, 6, 5),
                "value_date": date(2026, 6, 5),
                "description": "Payment",
                "bank_reference": "CHQ99",
                "debit": 3200.50,
                "credit": None,
                "balance_after": 45000.00,
                "raw_json": {"Channel": "Branch", "Detail": "Supplier payment"},
                "original_filename": "KTB6184_sample.xls",
                "source_row_number": 1,
                "match_status": "ignored",
                "match_reason": "โอนภายใน",
                "match_notes": "โอนภายใน — ไม่ใช่ค่าใช้จ่าย",
                "matched_ref_type": "internal_transfer",
                "matched_ref_id": "X4759",
                "match_confidence": 1.0,
            },
        ]
    )
    enriched = enrich_statement_rows(rows)
    sheets = build_account_sheets(enriched)
    return export_workbook(sheets, out_path, year=year, month=month)


def run_bank_statement_report(
    *,
    year: int | None = None,
    month: int | None = None,
    db_url: str | None = None,
    out_path: Path | None = None,
    verbose: bool = True,
) -> Path:
    """Load DB lines (with live match fields), enrich, and write monthly Excel."""
    import psycopg2

    if year is None or month is None:
        ry, rm = reporting_year_month()
        year = year if year is not None else ry
        month = month if month is not None else rm

    if not (1 <= int(month) <= 12):
        raise ValueError(f"Invalid month: {month}")

    url = db_url or supabase_db_url()
    if verbose:
        print(f"[bank-statement-report] year={year} month={month:02d}")

    conn = psycopg2.connect(url)
    try:
        df = load_statement_lines(conn, int(year), int(month))
    finally:
        conn.close()

    if verbose:
        print(f"[bank-statement-report] statement_lines={len(df):,}")

    if df.empty:
        raise RuntimeError(
            f"No bank.statement_lines for {year}-{int(month):02d}. "
            "Import statements first, or pass --year/--month for a month with data."
        )

    enriched = enrich_statement_rows(df)
    if verbose and "สถานะจับคู่" in enriched.columns:
        counts = enriched["สถานะจับคู่"].value_counts(dropna=False).to_dict()
        print(f"[bank-statement-report] match_status={counts}")

    sheets = build_account_sheets(enriched)
    if verbose:
        print(f"[bank-statement-report] sheets={list(sheets)}")

    dest = Path(out_path) if out_path else output_path(int(year), int(month))
    saved = export_workbook(sheets, dest, year=int(year), month=int(month))
    if verbose:
        print(f"[bank-statement-report] wrote {saved}")
    return saved
